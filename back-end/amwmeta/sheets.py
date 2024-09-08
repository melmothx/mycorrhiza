import csv
import logging
import re
import pprint
import sys
import xlrd
import openpyxl

logger = logging.getLogger(__name__)

pp = pprint.PrettyPrinter(indent=2)

def sheet_definitions():
    def add_name(name, value):
        if value:
            return "{} {}".format(name, value)
        else:
            return None
    def round_to_int(name, value):
        if value:
            if isinstance(value, float):
                return str(int(value))
            else:
                return str(value)
        else:
            return None

    definitions = {
        "calibre": {
            "type": "csv",
            "encoding": "utf-8-sig",
            "csv": {
                "delimiter": ",",
                "doublequote": True,
            },
            "mapping": [
                ('title', 'title'),
                ('date', 'pubdate'),
                ('publisher', 'publisher'),
                ('description', 'comments'),
                ('isbn', 'isbn'),
                ('language', 'languages', re.compile(r'\s*,\s*')),
                ('creator', 'authors', re.compile(r'\s*&\s*')),
                ('subject', 'tags', re.compile(r'\s*,\s*')),
                ('identifier', 'uuid'),
                ('identifier', 'id'),
            ]
        },
        "abebooks_home_base": {
            "type": "csv",
            "encoding": "latin-1",
            "csv": {
                "delimiter": ",",
                "doublequote": True,
            },
            "mapping": [
                ('title', 'Title'),
                ('date', 'Publisher Year'),
                ('publisher', 'publisher'),
                ('description', 'Description'),
                ('description', 'Synopsis'),
                ('place_date_of_publication_distribution', 'Publisher Place'),
                ('isbn', 'Isbn'),
                ('language', 'Language', re.compile(r'\s*,\s*')),
                ('creator', 'Author', re.compile(r'\s*/\s*')),
                ('subject', 'Keywords'),
                ('identifier', 'Book ID'),
            ],
        },
        "disordine": {
            "type": "excel",
            "start_at": 2,
            "mapping": [
                ('title', 'Titolo'),
                ('creator', 'Autore', re.compile(r'\s*/\s*')),
                ('publisher', 'Editore', re.compile(r'\s*/\s*')),
                ('subject', 'Serie'),
                ('language', 'Lingua'),
                ('subject', 'Argomento'),
                ('description', 'Note 1'),
                ('description', 'Note 2'),
                ('identifier', 'Posizione'),
                ('date', 'Data', None, (round_to_int,)),
                ('place_date_of_publication_distribution', 'Luogo editore'),
                ('country_of_publishing', 'Paese pubblicaz.'),
                ('description', 'Curatore', None, (add_name,)),
                ('description', 'Traduzione', None, (add_name,)),
                ('physical_description', 'Npag', None, (round_to_int, add_name)),
            ],
        },
    }
    return definitions

def parse_sheet(csv_type, sheet, **options):
    definitions = sheet_definitions()
    try:
        args = definitions[csv_type]
    except KeyError:
        return { "error": "Wrong type", "sample": {}, "number_of_records": 0 }
    sheet_error = None
    records = []
    if args['type'] == "csv":
        with open(sheet, newline='', encoding=args[csv_type]['encoding']) as csvfile:
            reader = csv.DictReader(csvfile, **args[csv_type]['csv'])
            try:
                for row in reader:
                    records.append(row)
            except UnicodeDecodeError as error:
                logger.error(error)
                sheet_error = str(error)

    elif args['type'] == "excel" and sheet.lower().endswith('.xls'):
        book = xlrd.open_workbook(sheet)
        for sh in book.sheets():
            columns = sh.ncols
            rows = sh.nrows
            if rows > 2:
                headers = []
                for cx in range(columns):
                    headers.append(sh.cell_value(0, cx))
                for row in range(args.get('start_at', 2), rows):
                    record = {}
                    for cx, cname in enumerate(headers):
                        record[cname] = sh.cell_value(row, cx)
                    records.append(record)

    elif args['type'] == "excel" and sheet.lower().endswith('.xlsx'):
        book = openpyxl.load_workbook(filename=sheet)
        for sh in book.worksheets:
            columns = sh.max_column
            rows = sh.max_row
            if rows > 2:
                headers = []
                for cx in range(1, columns + 1):
                    headers.append(sh.cell(1, cx).value)
                for row in range(args.get('start_at', 2) + 1, rows + 1):
                    record = {}
                    for cx, cname in enumerate(headers):
                        record[cname] = sh.cell(row, cx + 1).value
                    records.append(record)
        book.close()
    else:
        sheet_error = "Invalid file"
    if options.get('sample'):
        number_of_records = len(records)
        sample = {}
        if records:
            sample = records[0]
        return { "error": sheet_error, "number_of_records": number_of_records, "sample": sample }
    else:
        return normalize_records(records, args['mapping'])

def normalize_records(records, mapping):
    # let it crash with missing type
    out = []
    specification = []
    # add the defaults
    for spec in mapping:
        dest = spec[0]
        orig = spec[1]
        try:
            split = spec[2]
        except IndexError:
            split = None
        try:
            interpolate = spec[3]
        except IndexError:
            interpolate = None
        specification.append((dest, orig, split, interpolate))

    for rec in records:
        normal = {}
        for spec in specification:
            dest, orig, split, interpolate = spec
            if not dest in normal:
                normal[dest] = []

            value = rec.get(orig)
            if value:
                if interpolate:
                    for funct in interpolate:
                        value = funct(orig, value)
                value = str(value)
                if split:
                    normal[dest].extend(split.split(str(value)))
                else:
                    normal[dest].append(value)
        out.append(normal)
    return out

if __name__ == "__main__":
    pp.pprint(parse_sheet(sys.argv[1], sys.argv[2]))
